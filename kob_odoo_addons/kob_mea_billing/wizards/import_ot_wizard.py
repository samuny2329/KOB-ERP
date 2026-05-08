# -*- coding: utf-8 -*-
"""OT Import Wizard.

Imports monthly Online-team overtime hours from the HR OT 69.xlsx export.

File structure (one sheet per month):
    Sheets:    JAN 2026, FEB2026, MAR2026, APR2026, MAY2026, ...
    Row 1:     formatting / merged header
    Row 2:     header row -> col 1 = "ลำดับ" (seq), col 2 = "แผนก" (department),
               col 3 = employee name, col 4..N = Excel-serial date columns
    Row 3+:    data — each row = one employee for that month, daily OT hours
               in date columns (decimal hours, blank = no OT that day)

Sheet → billing-month mapping uses the sheet name (JAN 2026 → 2026-01-01).
The HR cycle is 21st→20th but is treated as covering the named calendar
month; this matches MEA bill ``billing_month`` (also a date set to day=1).

Online-team filter: department contains any of:
    'Online', 'Admin online', 'Online officer', 'Senior online'
(case-insensitive). 'Offline' is excluded explicitly.
"""
import base64
import io
import logging
import re
from collections import defaultdict
from datetime import date

from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Department keyword filter — case-insensitive substring match.
ONLINE_KEYWORDS = ("admin online", "online officer", "senior online", "online")
EXCLUDE_KEYWORDS = ("offline",)

THAI_MONTH_TO_NUM = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}


class MeaOtImportWizard(models.TransientModel):
    _name = "mea.ot.import.wizard"
    _description = "OT Import Wizard (HR xlsx → bill history)"

    xlsx_file = fields.Binary(string="OT 69.xlsx", required=True)
    xlsx_filename = fields.Char()
    meter_id = fields.Many2one(
        "mea.meter",
        string="Meter",
        required=True,
        help="MEA meter representing the site where the Online team operates "
             "(KK-16 by default). OT hours will be written to bill records "
             "of this meter.",
    )
    overwrite = fields.Boolean(
        string="Overwrite Existing",
        default=True,
    )

    summary = fields.Text(readonly=True)

    @api.model
    def default_get(self, fields_list):
        vals = super().default_get(fields_list)
        if "meter_id" in fields_list and not vals.get("meter_id"):
            kk16 = self.env["mea.meter"].search(
                [("meter_id", "=", "96746654")], limit=1
            )
            if kk16:
                vals["meter_id"] = kk16.id
        return vals

    # ---------- Actions ----------
    def action_import(self):
        self.ensure_one()
        if not self.xlsx_file:
            raise UserError("Please attach OT 69.xlsx.")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise UserError(
                "openpyxl is required. Install with: "
                "docker exec kob-odoo-19 pip install openpyxl"
            ) from exc

        raw = base64.b64decode(self.xlsx_file)
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise UserError(f"Could not read xlsx: {exc}") from exc

        per_month = {}  # date(y,m,1) -> {"hours": float, "employees": set, "raw_dept": set}

        for sheet_name in wb.sheetnames:
            month_d = self._sheet_to_month(sheet_name)
            if not month_d:
                _logger.info("Skipping sheet %s (no month parsed)", sheet_name)
                continue
            ws = wb[sheet_name]
            agg = self._aggregate_sheet(ws)
            if agg["hours"] > 0:
                per_month[month_d] = agg

        if not per_month:
            raise UserError(
                "No Online-team OT data found. Sheets scanned: "
                f"{', '.join(wb.sheetnames)}"
            )

        Bill = self.env["mea.bill.history"]
        updated, missing = [], []
        for month_d in sorted(per_month):
            agg = per_month[month_d]
            bill = Bill.search([
                ("meter_id", "=", self.meter_id.id),
                ("billing_month", "=", month_d),
            ], limit=1)
            if not bill:
                missing.append((month_d, agg))
                continue
            vals = {}
            if self.overwrite or not bill.ot_hours_online:
                vals["ot_hours_online"] = round(agg["hours"], 2)
            if self.overwrite or not bill.ot_employee_count:
                vals["ot_employee_count"] = len(agg["employees"])
            if vals:
                bill.write(vals)
                updated.append((month_d, agg, bill))

        lines = [
            f"Sheets scanned: {len(wb.sheetnames)}",
            f"Months with Online-team OT: {len(per_month)}",
            f"Bill records updated: {len(updated)}",
            f"Bill records missing: {len(missing)}",
            "",
            "Updated:",
        ]
        for month_d, agg, bill in updated:
            depts = ", ".join(sorted(agg["raw_dept"]))
            lines.append(
                f"  {month_d:%Y-%m}  hrs={agg['hours']:>8.1f}  "
                f"emp={len(agg['employees']):>3}  "
                f"depts=[{depts}]  → {bill.display_name}"
            )
        if missing:
            lines.append("")
            lines.append("Missing (no bill row exists):")
            for month_d, agg in missing:
                lines.append(
                    f"  {month_d:%Y-%m}  hrs={agg['hours']:>8.1f}  "
                    f"emp={len(agg['employees']):>3}"
                )

        self.summary = "\n".join(lines)
        return {
            "type": "ir.actions.act_window",
            "res_model": "mea.ot.import.wizard",
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
        }

    # ---------- Helpers ----------
    @staticmethod
    def _sheet_to_month(sheet_name):
        """Parse 'JAN 2026' / 'FEB2026' / 'MAR2026' → date(2026, M, 1)."""
        m = re.match(r"\s*([A-Z]{3})\s*(\d{4})\s*", sheet_name.upper())
        if not m:
            return None
        mon = THAI_MONTH_TO_NUM.get(m.group(1))
        if not mon:
            return None
        return date(int(m.group(2)), mon, 1)

    def _aggregate_sheet(self, ws):
        """Sum OT hours from rows where แผนก matches Online keywords."""
        hours = 0.0
        employees = set()
        raw_dept = set()
        # Header row is row 2 (1-indexed); data starts row 3.
        # Col 2 (B) = แผนก, col 3 (C) = employee, col 4+ = daily OT hours.
        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 4:
                continue
            dept = (row[1] or "").strip() if isinstance(row[1], str) else ""
            if not dept:
                continue
            dept_lc = dept.lower()
            if any(k in dept_lc for k in EXCLUDE_KEYWORDS):
                continue
            if not any(k in dept_lc for k in ONLINE_KEYWORDS):
                continue
            employee = (row[2] or "").strip() if isinstance(row[2], str) else ""
            row_hours = 0.0
            for cell in row[3:]:
                if cell is None:
                    continue
                try:
                    row_hours += float(cell)
                except (TypeError, ValueError):
                    continue
            if row_hours > 0:
                hours += row_hours
                if employee:
                    employees.add(employee)
                raw_dept.add(dept)
        return {"hours": hours, "employees": employees, "raw_dept": raw_dept}
