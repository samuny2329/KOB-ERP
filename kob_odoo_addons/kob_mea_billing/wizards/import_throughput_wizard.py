# -*- coding: utf-8 -*-
"""Throughput Import Wizard.

Imports monthly online-warehouse delivery quantities from an Odoo
``stock.move`` export (.xlsx) and writes them into matching
``mea.bill.history`` rows (matched by meter + billing_month).

Target Operation Types (configurable but defaults match KK-16 hub):
    - "BTV-WH2 (Online): Delivery Orders"
    - "KOB-WH2 (Online): Delivery Orders"

Output: per-month sums populated into ``order_qty_btv_wh2`` and
``order_qty_kob_wh2`` on existing bill records. Records whose
billing_month does not yet exist are reported as "missing"; the wizard
does not auto-create bill rows (energy data must come first).
"""
import base64
import io
import logging
from collections import defaultdict
from datetime import date

from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

DEFAULT_OPERATION_TYPES = (
    "BTV-WH2 (Online): Delivery Orders",
    "KOB-WH2 (Online): Delivery Orders",
)

# Column header → canonical key. Headers are case-sensitive in the Odoo export.
HEADER_MAP = {
    "Date Scheduled": "date",
    "Quantity": "qty",
    "Operation Type": "op_type",
}


class MeaThroughputImportWizard(models.TransientModel):
    _name = "mea.throughput.import.wizard"
    _description = "Throughput Import Wizard (stock.move xlsx → bill history)"

    xlsx_file = fields.Binary(string="Stock Move Export (.xlsx)", required=True)
    xlsx_filename = fields.Char()
    meter_id = fields.Many2one(
        "mea.meter",
        string="Meter",
        required=True,
        help="MEA meter representing the physical site that hosts these "
             "warehouses. Defaults to KK-16 (96746654) when present.",
    )
    overwrite = fields.Boolean(
        string="Overwrite Existing",
        default=True,
        help="When checked, replace existing order_qty_* values on matched "
             "bill records. When unchecked, only fill empty values.",
    )

    summary = fields.Text(readonly=True)

    @api.model
    def default_get(self, fields_list):
        vals = super().default_get(fields_list)
        if "meter_id" in fields_list and not vals.get("meter_id"):
            kk16 = self.env["mea.meter"].search(
                [("meter_id", "=", "96746654")], limit=1
            )
            if kk16:
                vals["meter_id"] = kk16.id
        return vals

    # ---------- Actions ----------
    def action_import(self):
        self.ensure_one()
        if not self.xlsx_file:
            raise UserError("Please attach a Stock Move xlsx export.")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise UserError(
                "openpyxl is required. Install with: "
                "docker exec kob-odoo-19 pip install openpyxl"
            ) from exc

        raw = base64.b64decode(self.xlsx_file)
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise UserError(f"Could not read xlsx: {exc}") from exc

        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration as exc:
            raise UserError("xlsx is empty.") from exc

        col_idx = {}
        for idx, name in enumerate(header):
            if name in HEADER_MAP:
                col_idx[HEADER_MAP[name]] = idx
        missing = {"date", "qty", "op_type"} - set(col_idx)
        if missing:
            raise UserError(
                f"Missing columns in xlsx: {sorted(missing)}. "
                f"Got headers: {list(header)}"
            )

        target_set = set(DEFAULT_OPERATION_TYPES)
        # Aggregate: (year, month, op_type) -> qty_sum
        agg = defaultdict(int)
        scanned = matched = 0
        for row in rows_iter:
            scanned += 1
            try:
                op_type = row[col_idx["op_type"]]
                if op_type not in target_set:
                    continue
                d_val = row[col_idx["date"]]
                q_val = row[col_idx["qty"]]
            except IndexError:
                continue
            if not d_val or q_val is None:
                continue
            month_d = self._coerce_month(d_val)
            if not month_d:
                continue
            try:
                qty_int = int(round(float(q_val)))
            except (TypeError, ValueError):
                continue
            agg[(month_d, op_type)] += qty_int
            matched += 1

        if not agg:
            raise UserError(
                f"No matching rows found. Scanned {scanned} rows. "
                f"Operation Type must be one of: {sorted(target_set)}."
            )

        # Pivot into per-month vals
        per_month = defaultdict(lambda: {"btv": 0, "kob": 0})
        for (month_d, op_type), qty in agg.items():
            if "BTV-WH2" in op_type:
                per_month[month_d]["btv"] += qty
            elif "KOB-WH2" in op_type:
                per_month[month_d]["kob"] += qty

        Bill = self.env["mea.bill.history"]
        updated_months, missing_months = [], []
        for month_d in sorted(per_month):
            qtys = per_month[month_d]
            bill = Bill.search([
                ("meter_id", "=", self.meter_id.id),
                ("billing_month", "=", month_d),
            ], limit=1)
            if not bill:
                missing_months.append((month_d, qtys))
                continue
            vals = {}
            if self.overwrite or not bill.order_qty_btv_wh2:
                vals["order_qty_btv_wh2"] = qtys["btv"]
            if self.overwrite or not bill.order_qty_kob_wh2:
                vals["order_qty_kob_wh2"] = qtys["kob"]
            if vals:
                bill.write(vals)
                updated_months.append((month_d, qtys, bill))

        lines = [
            f"Scanned rows: {scanned}",
            f"Matched rows (target ops): {matched}",
            f"Months in xlsx: {len(per_month)}",
            f"Bill records updated: {len(updated_months)}",
            f"Bill records missing: {len(missing_months)}",
            "",
            "Updated:",
        ]
        for month_d, qtys, bill in updated_months:
            lines.append(
                f"  {month_d:%Y-%m}  "
                f"BTV={qtys['btv']:>8,}  KOB={qtys['kob']:>8,}  "
                f"total={qtys['btv'] + qtys['kob']:>8,}  "
                f"→ {bill.display_name}"
            )
        if missing_months:
            lines.append("")
            lines.append("Missing (no bill row exists — create the bill first):")
            for month_d, qtys in missing_months:
                lines.append(
                    f"  {month_d:%Y-%m}  "
                    f"BTV={qtys['btv']:>8,}  KOB={qtys['kob']:>8,}"
                )

        self.summary = "\n".join(lines)
        return {
            "type": "ir.actions.act_window",
            "res_model": "mea.throughput.import.wizard",
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
        }

    @staticmethod
    def _coerce_month(value):
        """Return first-of-month date for a datetime/date/string cell value."""
        if isinstance(value, date):
            return value.replace(day=1)
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    from datetime import datetime
                    return datetime.strptime(value[:19], fmt).date().replace(day=1)
                except ValueError:
                    continue
        return None
