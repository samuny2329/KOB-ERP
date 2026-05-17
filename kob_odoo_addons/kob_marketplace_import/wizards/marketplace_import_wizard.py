# -*- coding: utf-8 -*-
"""Marketplace order import wizard.

Reads an Excel/CSV with one row per (order_sn, sku, qty) and creates
sale.order + lines + stock.picking with the x_kob_* fields the
Print_Label-App pipeline expects.

Excel layout (any of the column-name variants are accepted, case-insensitive):
    order_sn      | order # | order number
    shop          | shop_name
    order_date    | order_at | date
    sku           | sku_code
    qty           | quantity
    brand         | x_kob_brand
    fake          | fake_order   (Y / N / True / False / 1 / 0)
    [optional] price
"""
from __future__ import annotations

import base64
import csv
import io
import logging
from datetime import datetime, timezone

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Map of platform → expected res.partner.name (must match Print_Label-App).
#
# `employee` is an INTERNAL platform — used for staff welfare / monthly
# birthday-month perk orders that flow through the same Excel import
# but should NOT be commingled with public marketplace channels in
# reporting (different P&L line, no platform fee, internal stock cost
# only). Detected from order_sn prefix `KOBSO`/`BTVSO`/`CMNSO`.
_PARTNER_BY_PLATFORM = {
    "shopee":   "ECOMMERCE : SHOPEE",
    "tiktok":   "ECOMMERCE : TIKTOK",
    "lazada":   "ECOMMERCE : LAZADA",
    "employee": "Employee",
}
_PRETTY = {
    "shopee": "Shopee", "tiktok": "TikTok", "lazada": "Lazada",
    "employee": "Employee",
}

# Aliases we accept for each logical column.  English-only — Shopee /
# Lazada / TikTok official exports all use English headers (``Order
# ID``, ``SKU``, ``Order Date``, ``Quantity``, ``Unit Price``,
# ``Net Sale``, ``Shipping Fee``, …).  Thai column names are
# intentionally NOT accepted (per user policy: keep imports machine-
# readable).  Compared case-insensitively after _normalize() collapses
# punctuation + whitespace.
_COL_ALIASES = {
    "order_sn": (
        "order_sn", "order_no", "order_number", "order_id", "orderid",
        "order#", "order", "ordernumber", "order_ref", "ordersn",
        "platform_order_id", "marketplace_order_id",
        "order_reference", "orderreference",
    ),
    "shop": (
        "shop", "shop_name", "shopname", "store", "store_name",
        "source",
    ),
    "order_date": (
        "order_date", "date", "order_at", "date_order", "ordered_at",
        "orderdate", "ordereddate",
    ),
    "sku": (
        "sku", "sku_code", "code", "product_code", "item_id", "itemid",
        "product_sku", "default_code", "skucode",
        "order_lines_product", "orderlinesproduct", "product",
    ),
    "qty": (
        "qty", "quantity", "product_uom_qty", "count", "num",
        "order_lines_quantity", "orderlinesquantity",
    ),
    "brand": (
        "brand", "x_kob_brand", "brand_name", "brandname",
    ),
    "fake": (
        "fake", "fake_order", "x_kob_fake_order", "is_fake", "test",
    ),
    "price": (
        "price", "unit_price", "price_unit", "selling_price",
        "unitprice", "sellingprice", "net_sale", "netsale",
        "order_lines_unit_price", "orderlinesunitprice",
    ),
    # Optional marketplace gross/customer-paid total for the order.
    # When present, the importer back-fills any rounding gap so the Odoo
    # SO amount_total matches the marketplace's own customer-paid figure.
    "order_total": (
        "order_total", "ordertotal", "grand_total", "grandtotal",
        "customer_paid", "customerpaid", "total_amount", "totalamount",
        "total", "amount", "amount_total", "amounttotal",
        "gross_total", "grosstotal", "order_amount", "orderamount",
        "net_amount", "netamount", "paid_amount", "paidamount",
    ),
    "warehouse": (
        "warehouse", "warehousename", "warehouse_name", "wh", "whname",
    ),
    # ── Aligned with the standard Odoo SO export (Book1/Book2.xlsx) ──
    "delivery_date": (
        "delivery_date", "deliverydate", "commitment_date",
        "commitmentdate", "promised_date", "promiseddate",
    ),
    "order_type": (
        "order_type", "ordertype", "type", "sale_order_type",
        "saleordertype",
    ),
    "customer": (
        "customer", "partner", "partner_name", "partnername",
        "client", "buyer",
    ),
    "uom": (
        "uom", "unit_of_measure", "unitofmeasure",
        "order_lines_unit_of_measure", "orderlinesunitofmeasure",
        "product_uom", "productuom",
    ),
    "sales_team": (
        "sales_team", "salesteam", "team", "crm_team", "crmteam",
        "team_id",
    ),
    "tags": (
        "tags", "tag", "tag_ids", "order_tags", "ordertags",
    ),
    "source_document": (
        "source_document", "sourcedocument", "origin",
    ),
    "external_id": (
        "id", "external_id", "externalid", "xmlid",
    ),
}


_SKU_BRACKET_RE = __import__("re").compile(r"\[([^\]]+)\]")


def _extract_sku(value):
    """Extract SKU from values like '[SCPS220] SKINOXY Pro Cleanser ...'.
    If no [...] bracket, return as-is (assumes plain SKU)."""
    if value is None:
        return ""
    s = str(value).strip()
    m = _SKU_BRACKET_RE.search(s)
    return m.group(1).strip() if m else s


def _detect_platform_from_shop(shop_name):
    """Return 'shopee' | 'lazada' | 'tiktok' | 'employee' if shop_name
    carries a known prefix (e.g. 'Shopee_DaengGiMeoRi', 'Employee_HR').
    Allows mixed-platform xlsx files."""
    if not shop_name:
        return None
    s = str(shop_name).strip().lower()
    if s.startswith("shopee_"):
        return "shopee"
    if s.startswith("lazada_"):
        return "lazada"
    if s.startswith("tiktok_") or s.startswith("tikt_"):
        return "tiktok"
    if s.startswith("employee_") or s == "employee":
        return "employee"
    return None


# Order-sn prefixes that indicate an internal employee welfare order.
# KOBSO = Kiss of Beauty Sales Order. BTVSO = Beauty Ville Sales Order.
# CMNSO = Cosmonation Sales Order. Numeric / marketplace order ids never
# start with these letters so the detection is unambiguous.
_EMPLOYEE_ORDER_SN_PREFIXES = ("KOBSO", "BTVSO", "CMNSO")


def _detect_platform_from_order_sn(order_sn):
    """Internal employee orders identified by SO-name prefix.

    Excel uploads for employee perks reuse the marketplace import wizard
    but the `shop` column is often blank or a placeholder like
    `Lazada_Unknown` — without this fallback those rows would route to
    the wrong partner (LAZADA) and pollute marketplace reports.
    """
    if not order_sn:
        return None
    s = str(order_sn).strip().upper()
    if s.startswith(_EMPLOYEE_ORDER_SN_PREFIXES):
        return "employee"
    return None


def _normalize(s):
    """Lower-case, strip leading/trailing space, and squeeze inner
    whitespace + common punctuation away so that ``Order #``,
    ``order_number``, ``Order ID`` and ``ORDER NUMBER`` all collapse
    to the same canonical form for alias matching.  Non-ASCII
    characters (incl. Thai) are dropped entirely so they cannot
    accidentally match."""
    if s is None:
        return ""
    out = str(s).strip().lower()
    keep = []
    for ch in out:
        # Keep ASCII alnum + underscore.  Drop everything else.
        if (ch.isascii() and ch.isalnum()) or ch == "_":
            keep.append(ch)
    return "".join(keep)


# Pre-normalise the alias table once at import-time so column matching
# can compare apples to apples.
_COL_ALIASES_NORM = {
    logical: tuple(_normalize(a) for a in aliases)
    for logical, aliases in _COL_ALIASES.items()
}


def _to_bool(v):
    if isinstance(v, bool):
        return v
    if v is None or v == "":
        return False
    s = str(v).strip().lower()
    return s in ("y", "yes", "true", "1", "t")


class MarketplaceImportWizard(models.TransientModel):
    _name = "kob.marketplace.import.wizard"
    _description = "Import Marketplace Orders"

    platform = fields.Selection(
        [("shopee", "Shopee"), ("tiktok", "TikTok"), ("lazada", "Lazada")],
        default="shopee",
        help="Fallback platform when the Source column in xlsx doesn't "
             "carry a 'Shopee_/Lazada_/Tiktok_' prefix.",
    )
    company_id = fields.Many2one(
        "res.company",
        default=lambda self: self.env.company,
        help="Fallback company when shop is not in wms.shop.company.map.",
    )
    warehouse_id = fields.Many2one(
        "stock.warehouse",
        domain="[('company_id', '=', company_id)]",
        compute="_compute_default_warehouse",
        store=True, readonly=False,
        help="Fallback warehouse when xlsx 'Warehouse' column is empty "
             "and shop mapping doesn't resolve.",
    )

    @api.depends("company_id", "platform")
    def _compute_default_warehouse(self):
        # Convention: marketplace orders default to the company's
        # "Online" warehouse — name contains "Online" — fallback to
        # the first warehouse of that company.
        for w in self:
            if not w.company_id:
                w.warehouse_id = False
                continue
            online = self.env["stock.warehouse"].search([
                ("company_id", "=", w.company_id.id),
                ("name", "ilike", "Online"),
            ], limit=1)
            if not online:
                online = self.env["stock.warehouse"].search([
                    ("company_id", "=", w.company_id.id),
                ], limit=1)
            w.warehouse_id = online.id if online else False
    file_data = fields.Binary("Excel/CSV File")
    filename = fields.Char()
    # Multi-file mode — drag-and-drop the whole folder at once.  Each
    # ir.attachment is one xlsx/csv file.  Platform is auto-detected
    # from the filename's first underscore-separated token (Shopee /
    # Lazada / TikTok).
    attachment_ids = fields.Many2many(
        "ir.attachment",
        string="Excel/CSV Files (batch)",
        help="Drop the whole folder here — each file is processed in "
             "turn.  Platform is auto-detected from the filename "
             "(Shopee_… / Lazada_… / Tiktok_…).",
    )
    auto_confirm = fields.Boolean("Confirm SOs", default=True)
    auto_assign = fields.Boolean("Reserve stock (assign)", default=True)
    confirm_create_missing = fields.Boolean(
        "I confirm auto-create the missing products listed above",
        default=False,
        help="Tick this BEFORE clicking Import when the preview shows "
             "missing SKUs. Confirms you reviewed each one and accept "
             "Excel name + price as the new product master record.",
    )
    auto_create_product = fields.Boolean(
        "Auto-create missing products",
        default=True,
        help="If a SKU in the file is not in product.product, create a "
             "lot-tracked storable product on the fly with default_code "
             "= SKU and x_kob_brand auto-derived from the filename.",
    )
    auto_adjust_lot = fields.Boolean(
        "Adjust stock (create lot + qty)",
        default=True,
        help="After auto-creating a product (or for any imported line "
             "where on-hand qty is insufficient), generate a stock.lot "
             "and post an inventory adjustment so the Sale Order can be "
             "reserved at confirmation time.  Lot name = "
             "AUTO-<YYYYMMDD>-<SKU>.",
    )
    fbs_auto_route = fields.Boolean(
        "FBS auto-route", default=True,
        help="If the order_sn ends with -FBS (Shopee Fulfilled-By-Shopee), "
             "route to the company's *-SHOPEE warehouse instead of *-Online.",
    )

    # Results
    state = fields.Selection(
        [
            ("draft", "Draft"),
            ("preview", "Preview"),
            ("running", "Running"),
            ("done", "Done"),
        ],
        default="draft",
    )
    preview_orders_count = fields.Integer(readonly=True, default=0)
    preview_lines_count = fields.Integer(readonly=True, default=0)
    preview_duplicates_count = fields.Integer(readonly=True, default=0)
    preview_missing_skus = fields.Text(readonly=True)
    preview_unmapped_shops = fields.Text(readonly=True)
    preview_companies_breakdown = fields.Text(readonly=True)
    preview_warehouses_breakdown = fields.Text(readonly=True)
    preview_can_proceed = fields.Boolean(readonly=True, default=False)
    preview_log = fields.Text(readonly=True)
    log = fields.Text(readonly=True)
    sale_order_ids = fields.Many2many("sale.order", readonly=True)
    duplicate_order_ids = fields.Many2many(
        "sale.order", "kob_mp_import_dup_rel", "wiz_id", "so_id",
        readonly=True, string="Duplicate Sale Orders",
        help="Existing SOs that matched an imported order_sn — skipped.",
    )
    progress_total = fields.Integer(readonly=True, default=0)
    progress_done = fields.Integer(readonly=True, default=0)
    progress_pct = fields.Integer(readonly=True, default=0,
                                   help="Percent of orders processed.")
    imported_count = fields.Integer(readonly=True, default=0,
                                     string="Imported")
    duplicate_count = fields.Integer(readonly=True, default=0,
                                      string="Duplicates Skipped")
    unresolved_count = fields.Integer(readonly=True, default=0,
                                       string="Unresolved Skipped")
    duplicate_summary = fields.Text(readonly=True,
                                     string="Duplicate Details",
                                     help="One line per skipped duplicate "
                                          "with the existing SO it matched.")

    # ── Parsing ────────────────────────────────────────────────────
    @staticmethod
    def _detect_platform(filename):
        """Return one of 'shopee'|'lazada'|'tiktok' based on the
        filename's first token, or None if undetectable."""
        if not filename:
            return None
        stem = filename.rsplit(".", 1)[0]
        first = stem.split(" ", 1)[0].split("_", 1)[0].lower()
        if first.startswith("shopee"):
            return "shopee"
        if first.startswith("lazada"):
            return "lazada"
        if first.startswith("tiktok") or first.startswith("tikt"):
            return "tiktok"
        return None

    @staticmethod
    def _derive_brand_shop(filename):
        """Pull brand/shop from filename pattern
        ``<Platform>_<Brand>[_<Word>...] <Date> <Shift>...``."""
        if not filename:
            return (None, None)
        stem = filename.rsplit(".", 1)[0]
        head = stem.split(" ", 1)[0]
        parts = head.split("_")
        if len(parts) >= 2:
            shop = "_".join(parts[1:])
            return (shop, shop)
        return (None, None)

    def _parse_blob(self, raw, filename):
        """Parse one (raw bytes, filename) pair into a list of
        record dicts.  Replaces the previous instance-bound
        ``_parse_file`` so we can iterate over multiple files in a
        single import run."""
        name = (filename or "").lower()
        derived_brand, derived_shop = self._derive_brand_shop(filename)

        if name.endswith(".csv"):
            text = raw.decode("utf-8-sig")
            rdr = csv.reader(io.StringIO(text))
            rows = list(rdr)
            if not rows:
                raise UserError(_("File is empty."))
            header = [_normalize(c) for c in rows[0]]
            data = rows[1:]
        else:
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise UserError(_(
                    "openpyxl is not installed in the Odoo container — "
                    "use a CSV file instead.",
                ))
            wb = load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb.active
            iter_ = ws.iter_rows(values_only=True)
            header_row = next(iter_, None)
            if not header_row:
                raise UserError(_("File is empty."))
            header = [_normalize(c) for c in header_row]
            data = [r for r in iter_ if any(c is not None for c in r)]

        # Resolve column indexes (case + punctuation insensitive).
        col = {}
        for logical, aliases in _COL_ALIASES_NORM.items():
            for i, h in enumerate(header):
                if h in aliases:
                    col[logical] = i
                    break

        for required in ("order_sn", "sku", "qty"):
            if required not in col:
                # Show both the accepted aliases and the columns we
                # actually saw — much easier to fix the file.
                raise UserError(_(
                    "Required column missing: %(name)s.\n\n"
                    "Accepted aliases (case-insensitive): %(aliases)s\n\n"
                    "Columns found in your file: %(found)s\n\n"
                    "Fix: rename one of your columns to any accepted "
                    "alias, or add a new alias to _COL_ALIASES in "
                    "kob_marketplace_import/wizards/"
                    "marketplace_import_wizard.py.",
                ) % {
                    "name": required,
                    "aliases": ", ".join(_COL_ALIASES[required]),
                    "found": ", ".join(repr(h) for h in header) or "(empty)",
                })

        records = []
        last_order_sn = None
        last_shop = None
        last_order_date = None
        last_delivery_date = None
        last_warehouse = None
        last_order_type = None
        last_customer = None
        last_sales_team = None
        last_tags = None
        last_source_document = None
        last_external_id = None
        last_order_total = None
        for r in data:
            order_sn = r[col["order_sn"]]
            if order_sn:
                last_order_sn = str(order_sn).strip()
                last_shop = (
                    str(r[col["shop"]]).strip()
                    if "shop" in col and r[col["shop"]]
                    else derived_shop
                )
                last_order_date = (
                    r[col["order_date"]]
                    if "order_date" in col else None
                )
                last_delivery_date = (
                    r[col["delivery_date"]]
                    if "delivery_date" in col else None
                )
                last_warehouse = (
                    str(r[col["warehouse"]]).strip()
                    if "warehouse" in col and r[col["warehouse"]]
                    else None
                )
                last_order_type = (
                    str(r[col["order_type"]]).strip()
                    if "order_type" in col and r[col["order_type"]]
                    else None
                )
                last_customer = (
                    str(r[col["customer"]]).strip()
                    if "customer" in col and r[col["customer"]]
                    else None
                )
                last_sales_team = (
                    str(r[col["sales_team"]]).strip()
                    if "sales_team" in col and r[col["sales_team"]]
                    else None
                )
                last_tags = (
                    str(r[col["tags"]]).strip()
                    if "tags" in col and r[col["tags"]]
                    else None
                )
                last_source_document = (
                    str(r[col["source_document"]]).strip()
                    if "source_document" in col and r[col["source_document"]]
                    else None
                )
                last_external_id = (
                    str(r[col["external_id"]]).strip()
                    if "external_id" in col and r[col["external_id"]]
                    else None
                )
                # Capture marketplace gross/customer-paid total on the order
                # header row (first row of each order group). Used later to
                # close any rounding gap between Odoo's computed amount_total
                # and the marketplace's reported customer-paid amount.
                if "order_total" in col and r[col["order_total"]] not in (None, ""):
                    try:
                        last_order_total = float(r[col["order_total"]])
                    except (TypeError, ValueError):
                        last_order_total = None
                else:
                    last_order_total = None
            else:
                if not last_order_sn:
                    continue
            sku_raw = r[col["sku"]] if "sku" in col else None
            if not sku_raw:
                continue
            shop_for_row = last_shop or derived_shop
            # Platform detection: shop prefix is primary; fallback to
            # order_sn prefix so internal employee orders (KOBSO/BTVSO/
            # CMNSO) don't get misrouted to a marketplace partner when
            # the Excel shop column is a placeholder like Lazada_Unknown.
            row_platform = (
                _detect_platform_from_order_sn(last_order_sn)
                or _detect_platform_from_shop(shop_for_row)
            )
            records.append({
                "order_sn": last_order_sn,
                "shop": shop_for_row,
                "order_date": last_order_date,
                "delivery_date": last_delivery_date,
                "order_type": last_order_type,
                "customer": last_customer,
                "sales_team": last_sales_team,
                "tags": last_tags,
                "source_document": last_source_document,
                "external_id": last_external_id,
                "_row_platform": row_platform,
                "_row_warehouse": last_warehouse,
                "sku": _extract_sku(sku_raw),
                "raw_name": str(sku_raw).strip() if sku_raw else "",
                "qty": float(r[col["qty"]] or 0),
                "uom": (
                    str(r[col["uom"]]).strip()
                    if "uom" in col and r[col["uom"]]
                    else None
                ),
                "brand": (
                    str(r[col["brand"]]).strip()
                    if "brand" in col and r[col["brand"]]
                    else derived_brand
                ),
                "fake": _to_bool(r[col["fake"]]) if "fake" in col else False,
                "price": float(r[col["price"]] or 0)
                          if "price" in col else 0.0,
                "order_total": last_order_total,
            })
        return records

    # ── Helpers ───────────────────────────────────────────────────
    def _get_partner(self):
        partner = self.env["res.partner"].search(
            [("name", "=", _PARTNER_BY_PLATFORM[self.platform])], limit=1,
        )
        if not partner:
            partner = self.env["res.partner"].create({
                "name": _PARTNER_BY_PLATFORM[self.platform],
                "is_company": True, "customer_rank": 1,
            })
        return partner

    def _get_source(self, shop_name):
        pretty = _PRETTY[self.platform]
        if shop_name.lower().startswith(f"{pretty.lower()}_"):
            full = shop_name
        else:
            full = f"{pretty}_{shop_name}"
        s = self.env["utm.source"].search([("name", "=", full)], limit=1)
        if not s:
            s = self.env["utm.source"].create({"name": full})
        return s

    def _resolve_product(self, sku, brand=None, excel_name=None,
                         excel_price=None):
        """Find the product.product matching ``sku`` (default_code OR
        x_kob_sku_code OR ``[SKU]`` prefix in name).  If ``self.
        auto_create_product`` is set and nothing matches, create a
        new product on the fly so the import can proceed.

        ``excel_name``/``excel_price`` (Excel "Order Lines/Product" and
        "Order Lines/Unit Price") are used to seed the new product so
        the auto-created record carries the marketplace-correct display
        name + list price 1:1 — no more placeholder "auto-imported".
        """
        # Company-agnostic search: deduped product master may have
        # ``product_template.company_id`` set on a single company while
        # the wizard runs under a different ``allowed_company_ids``
        # scope (e.g. importing a company-2 order while canonical
        # template lives in company-4). Without an explicit override
        # the Odoo MultiCompany rule filters the canonical out, the
        # wizard falls through to auto-create, and a brand-new
        # duplicate gets written — which Phase C's BEFORE INSERT
        # trigger then blocks, aborting the whole import transaction.
        #
        # Fix: search as sudo() with allowed_company_ids covering every
        # company so the canonical record is always visible. The
        # company-aware stock check happens later in the multi_company
        # override (`_kob_product_has_stock`), so this only affects
        # *discovery*, not the per-company stock gating that the user
        # already relies on.
        all_company_ids = self.env['res.company'].sudo().search([]).ids
        ProductPP = self.env["product.product"].sudo().with_context(
            allowed_company_ids=all_company_ids,
        )
        prod = ProductPP.search([("default_code", "=", sku)], limit=1)
        if not prod:
            prod = ProductPP.search([("x_kob_sku_code", "=", sku)], limit=1)
        if not prod:
            # Match by [SKU] prefix in name
            prod = ProductPP.search(
                [("name", "=ilike", f"[{sku}]%")], limit=1,
            )
        if prod:
            return prod
        if not self.auto_create_product:
            return prod
        # Auto-create — storable + lot-tracked so the WMS pipeline can
        # reserve / pick / pack.  Stock is seeded later via the
        # ``_ensure_stock_with_lot`` helper triggered for each order
        # line.
        #
        # Strip the leading "[SKU]" prefix from the Excel name (Odoo's
        # display_name will add it back automatically).
        import re as _re
        name = ""
        if excel_name:
            name = _re.sub(r"^\[[^\]]+\]\s*", "", str(excel_name)).strip()
        if not name:
            name = ("%s — auto-imported" % brand) if brand else "auto-imported"
        # Odoo 19 split product.template.type — 'product' is no longer
        # accepted.  We use type='consu' + is_storable=True for the
        # stockable + lot-tracked variant.
        new_prod_vals = {
            "name":           name,
            "default_code":   sku,
            "barcode":        sku,
            "x_kob_sku_code": sku,
            "x_kob_brand":    brand or "",
            "type":           "consu",
            "tracking":       "lot",
            "sale_ok":        True,
            "purchase_ok":    False,
        }
        if excel_price and float(excel_price) > 0:
            new_prod_vals["list_price"] = round(float(excel_price), 2)
            new_prod_vals["standard_price"] = round(float(excel_price), 2)
        if "is_storable" in self.env["product.template"]._fields:
            new_prod_vals["is_storable"] = True
        # ``barcode`` is unique on product.product; in the rare case
        # the same SKU got registered as a barcode by another import
        # path, fall back to no-barcode rather than crashing.
        try:
            new_prod = self.env["product.product"].sudo().create(new_prod_vals)
        except Exception as e:
            _logger.warning(
                "Barcode '%s' clashes with an existing product — "
                "creating without barcode (%s)", sku, e,
            )
            new_prod_vals.pop("barcode", None)
            new_prod = self.env["product.product"].sudo().create(new_prod_vals)
        _logger.info(
            "Auto-created lot-tracked product %s (%s) for marketplace import",
            new_prod.id, sku,
        )
        return new_prod

    def _ensure_stock_with_lot(self, product, qty, warehouse, unit_cost=0.0,
                               order_sn=None):
        """Create a stock.lot for ``product`` (if needed) and post an
        inventory adjustment so ``qty`` is available in
        ``warehouse``'s stock location.  Returns the lot.

        ``unit_cost`` is stored on stock.lot.x_kob_cost_per_unit so
        per-lot valuation reports get the right basis.

        Idempotent: if a lot named AUTO-<YYYYMMDD>-<SKU> already
        exists with enough quantity, just top it up.
        """
        from datetime import date as _date

        if not product or qty <= 0 or not warehouse:
            return self.env["stock.lot"]

        # Pick the location the outgoing picking_type actually reserves
        # from — on KOB-WH2 (Online) that is K-On/Stock/PICKFACE (511),
        # NOT the parent K-On/Stock (5). Falling back to the parent
        # lot_stock_id is fine for warehouses without a pickface child.
        out_pt = self.env["stock.picking.type"].search([
            ("warehouse_id", "=", warehouse.id),
            ("code", "=", "outgoing"),
        ], limit=1, order="sequence asc, id asc")
        loc = (
            out_pt.default_location_src_id
            if out_pt and out_pt.default_location_src_id
            else warehouse.lot_stock_id
        )
        if not loc:
            return self.env["stock.lot"]

        # Upgrade product to lot-tracked so the lot-bound quant write
        # below is accepted by Odoo's stock engine. Existing untracked
        # stock stays as a single legacy quant (lot_id=NULL); new
        # marketplace lots layer on top.
        if product.tracking == 'none':
            try:
                product.sudo().write({"tracking": "lot"})
            except Exception:
                pass  # may fail if product has active reservations

        sku = product.default_code or str(product.id)
        # Lot name traceability: prefer the marketplace order # so each
        # imported order's stock has a clean audit trail back to the
        # platform's order_sn.  Fall back to AUTO-<date>-<sku> when no
        # order_sn is available (e.g. manual top-ups).
        if order_sn:
            lot_name = "%s-%s" % (str(order_sn).strip(), sku)
        else:
            lot_name = "AUTO-%s-%s" % (_date.today().strftime("%Y%m%d"), sku)
        Lot = self.env["stock.lot"].sudo()
        lot = Lot.search([
            ("name", "=", lot_name),
            ("product_id", "=", product.id),
        ], limit=1)
        if not lot:
            lot = Lot.create({
                "name":               lot_name,
                "product_id":         product.id,
                "company_id":         warehouse.company_id.id,
                "x_kob_cost_per_unit": (
                    unit_cost or float(product.standard_price or 0)
                ),
            })
        elif unit_cost and not lot.x_kob_cost_per_unit:
            # First time we're seeing a meaningful cost — record it.
            lot.x_kob_cost_per_unit = unit_cost

        # Top up the lot quantity at the warehouse stock location via
        # Odoo's inventory adjustment API:
        #   1. set inventory_quantity (the "counted" qty)
        #   2. call _apply_inventory() to commit it as a move
        # kob_wms's override now accepts the date kwarg so this is safe.
        Quant = self.env["stock.quant"].sudo().with_context(
            inventory_mode=True,
        )
        quant = Quant.search([
            ("product_id",  "=", product.id),
            ("location_id", "=", loc.id),
            ("lot_id",      "=", lot.id),
        ], limit=1)
        target_qty = float(quant.quantity or 0) + qty if quant else qty
        if not quant:
            quant = Quant.create({
                "product_id":  product.id,
                "location_id": loc.id,
                "lot_id":      lot.id,
                "inventory_quantity": target_qty,
            })
        else:
            quant.inventory_quantity = target_qty
        try:
            quant._apply_inventory()
        except Exception:
            # Fallback: write quantity directly (legacy path)
            quant.write({"quantity": target_qty})
        return lot

    def _kob_preview_anomalies(self, by_order, records, resolve_company_fn,
                               Warehouse):
        """Scan parsed Excel BEFORE creating any SO and return a list
        of anomalies with severity (BLOCK / WARN / INFO).

        Catches the patterns that bit the user in past runs:
          • Cross-company team mismatch (BTV-routed order with KOB team)
          • Missing required cell (order_sn / sku / qty)
          • Unparseable date (Order Date / Delivery Date)
          • Tag label not yet in crm.tag (will be auto-created)
          • Sales Team / Order Type label not in master
          • Stock insufficient at target warehouse's outgoing location
          • Duplicate within the same file (same order_sn + same SKU twice)
        """
        anomalies = []
        Tag = self.env["crm.tag"].sudo()
        Team = self.env["crm.team"].sudo()
        OrderType = self.env["sale.order.type"].sudo()
        Product = self.env["product.product"].sudo()
        Quant = self.env["stock.quant"].sudo()
        PickType = self.env["stock.picking.type"].sudo()
        seen_dup_keys = set()

        # Pre-load known tag/team/order-type labels for quick membership tests
        known_tag_labels = {(t.name or "").strip().lower(): t for t in Tag.search([])}
        known_team_labels = {(t.name or "").strip().lower(): t for t in Team.search([])}
        known_ot_labels = {(t.name or "").strip().lower(): t for t in OrderType.search([])}

        for r in records:
            sn = (r.get("order_sn") or "").strip()
            sku = (r.get("sku") or "").strip()
            qty = r.get("qty") or 0

            # — Required cells —
            if not sn:
                anomalies.append({
                    "severity": "BLOCK",
                    "msg": f"Row in {r.get('__source_file','?')} has empty Order Reference",
                })
                continue
            if not sku:
                anomalies.append({
                    "severity": "BLOCK",
                    "msg": f"{sn} — missing SKU (Order Lines/Product) column value",
                })
            if not qty or qty <= 0:
                anomalies.append({
                    "severity": "BLOCK",
                    "msg": f"{sn} — quantity is {qty} (must be > 0)",
                })

            # — Duplicate within file (same order_sn + SKU) —
            dup_key = (sn, sku)
            if sku and dup_key in seen_dup_keys:
                anomalies.append({
                    "severity": "WARN",
                    "msg": f"{sn} — SKU '{sku}' appears twice in the same order",
                })
            seen_dup_keys.add(dup_key)

            # — Date parseability —
            for fname, key in [("Order Date", "order_date"),
                               ("Delivery Date", "delivery_date")]:
                v = r.get(key)
                if v:
                    try:
                        self._normalize_date(v)
                    except Exception:
                        anomalies.append({
                            "severity": "WARN",
                            "msg": f"{sn} — {fname} '{v!r}' is not a valid date",
                        })

            # — Tag labels new vs existing (INFO, not blocking) —
            if r.get("tags"):
                import re as _re
                for label in _re.split(r"[,;|]", str(r["tags"])):
                    label = label.strip()
                    if label and label.lower() not in known_tag_labels:
                        anomalies.append({
                            "severity": "INFO",
                            "msg": f"{sn} — Tag '{label}' will be auto-created in crm.tag",
                        })
                        # add to known set so we don't repeat per row
                        known_tag_labels[label.lower()] = True

            # — Sales Team / Order Type labels —
            if r.get("sales_team"):
                if r["sales_team"].strip().lower() not in known_team_labels:
                    anomalies.append({
                        "severity": "INFO",
                        "msg": f"{sn} — Sales Team '{r['sales_team']}' will be auto-created",
                    })
                    known_team_labels[r["sales_team"].strip().lower()] = True
            if r.get("order_type"):
                if r["order_type"].strip().lower() not in known_ot_labels:
                    anomalies.append({
                        "severity": "WARN",
                        "msg": f"{sn} — Order Type '{r['order_type']}' not in master (will fallback to Normal Order)",
                    })

        # — Per-order checks (cross-company, stock availability) —
        for (platform, order_sn), lines in by_order.items():
            head = lines[0]
            shop = head.get("shop") or ""
            target_co = resolve_company_fn(platform, shop)
            row_wh_name = head.get("_row_warehouse")
            wh = Warehouse.browse()
            if row_wh_name:
                explicit = Warehouse.search(
                    [("name", "=", row_wh_name)], limit=1,
                )
                if explicit:
                    wh = explicit
                    target_co = explicit.company_id
            if not wh:
                wh = Warehouse.search([
                    ("company_id", "=", target_co.id),
                ], limit=1) or self.warehouse_id

            # — Cross-company team check —
            if head.get("sales_team"):
                team = Team.search(
                    [("name", "=ilike", head["sales_team"])], limit=1,
                )
                if team and team.company_id and team.company_id != target_co:
                    anomalies.append({
                        "severity": "INFO",
                        "msg": (f"{order_sn} routes to {target_co.name} but team "
                                f"'{head['sales_team']}' belongs to "
                                f"{team.company_id.name} — will auto-create matching team"),
                    })

            # — Stock availability prediction —
            out_pt = PickType.search([
                ("warehouse_id", "=", wh.id),
                ("code", "=", "outgoing"),
            ], limit=1, order="sequence asc, id asc")
            src_loc = (out_pt.default_location_src_id
                       if out_pt and out_pt.default_location_src_id
                       else wh.lot_stock_id)
            if src_loc:
                for ln in lines:
                    sku = (ln.get("sku") or "").strip()
                    need = float(ln.get("qty") or 0)
                    if not sku or need <= 0:
                        continue
                    prod = Product.search([
                        "|", "|",
                        ("default_code", "=", sku),
                        ("x_kob_sku_code", "=", sku),
                        ("name", "=ilike", "[%s]%%" % sku),
                    ], limit=1)
                    if not prod:
                        continue  # missing_skus already covered
                    on_hand = sum(Quant.search([
                        ("product_id", "=", prod.id),
                        ("location_id", "=", src_loc.id),
                    ]).mapped("quantity"))
                    if on_hand < need and not self.auto_adjust_lot:
                        anomalies.append({
                            "severity": "WARN",
                            "msg": (f"{order_sn} — SKU '{sku}' needs {need} but only "
                                    f"{on_hand} on hand at {src_loc.name} "
                                    "(enable Adjust stock to auto-top-up)"),
                        })

        return anomalies

    # PostgreSQL advisory-lock key for serialising end-of-import stamps
    # across concurrent import sessions (cron-polled marketplaces fire
    # many wizards in parallel — without this they race on the same
    # sale_order rows and one session loses the whole stamp to a
    # SerializationFailure). Picked as a stable arbitrary 64-bit int.
    _KOB_FINAL_STAMP_LOCK_KEY = 7438291023641

    def _kob_final_stamp(self, records):
        """End-of-import safety net: re-write Excel-derived fields on
        every SO via raw SQL.

        Reads `records` (the parsed Excel rows), groups by order_sn,
        looks up the SO created for that order, then issues:
          UPDATE sale_order SET commitment_date=, origin=, date_order=
          WHERE id=X
        plus DELETE+INSERT on sale_order_tag_rel to match the row's
        "Tags" column 1:1.

        Bypasses ORM so compute methods / onchange chains can't wipe the
        values mid-flow. Idempotent — safe to run multiple times.

        Concurrency: takes a transaction-scoped advisory lock so two
        concurrent imports can't race on the same rows. Each per-SO
        write is wrapped in a SAVEPOINT — a SerializationFailure on one
        SO logs + skips that SO instead of rolling back the whole stamp.
        """
        import re as _re
        if not records:
            return

        # Serialise concurrent imports — released at txn commit/rollback
        self.env.cr.execute(
            "SELECT pg_advisory_xact_lock(%s)",
            (self._KOB_FINAL_STAMP_LOCK_KEY,),
        )

        # Group records by order_sn — preserve first occurrence as
        # canonical (Excel header row) for scalar fields, accumulate
        # tags across all lines.
        by_sn = {}
        for r in records:
            sn = (r.get("order_sn") or "").strip()
            if not sn:
                continue
            slot = by_sn.setdefault(sn, {
                "order_date": None,
                "delivery_date": None,
                "origin": None,
                "tag_labels": set(),
            })
            if r.get("order_date") and slot["order_date"] is None:
                slot["order_date"] = r["order_date"]
            if r.get("delivery_date") and slot["delivery_date"] is None:
                slot["delivery_date"] = r["delivery_date"]
            if r.get("source_document") and slot["origin"] is None:
                slot["origin"] = str(r["source_document"]).strip()
            if r.get("tags"):
                for label in _re.split(r"[,;|]", str(r["tags"])):
                    label = label.strip()
                    if label:
                        slot["tag_labels"].add(label)

        # Pre-resolve all tag labels into crm.tag ids (create if missing)
        all_labels = set()
        for slot in by_sn.values():
            all_labels.update(slot["tag_labels"])
        tag_label_to_id = {}
        for label in all_labels:
            tag = self.env["crm.tag"].search(
                [("name", "=ilike", label)], limit=1,
            )
            if not tag:
                tag = self.env["crm.tag"].sudo().create({"name": label})
            tag_label_to_id[label] = tag.id

        # Bulk UPDATE per SO. Each row gets its own SAVEPOINT so a
        # transient serialization failure on one SO logs + skips that
        # SO instead of poisoning the whole transaction.
        SaleOrder = self.env["sale.order"]
        updated = 0
        tagged = 0
        skipped_concurrent = 0
        for sn, slot in by_sn.items():
            so = SaleOrder.search([("name", "=", sn)], limit=1) \
                or SaleOrder.search([("client_order_ref", "=", sn)], limit=1)
            if not so:
                continue
            sp_name = f"kob_stamp_{so.id}"
            try:
                self.env.cr.execute(f"SAVEPOINT {sp_name}")
                sets, args = [], []
                if slot["order_date"]:
                    sets.append("date_order=%s")
                    args.append(self._normalize_date(slot["order_date"]))
                if slot["delivery_date"]:
                    sets.append("commitment_date=%s")
                    args.append(self._normalize_date(slot["delivery_date"]))
                if slot["origin"]:
                    sets.append("origin=%s")
                    args.append(slot["origin"])
                if sets:
                    args.append(so.id)
                    self.env.cr.execute(
                        "UPDATE sale_order SET " + ", ".join(sets)
                        + " WHERE id=%s",
                        args,
                    )
                    updated += 1
                if slot["tag_labels"]:
                    tag_ids = [tag_label_to_id[lab] for lab in slot["tag_labels"]
                               if lab in tag_label_to_id]
                    if tag_ids:
                        self.env.cr.execute(
                            "DELETE FROM sale_order_tag_rel WHERE order_id=%s",
                            (so.id,),
                        )
                        for tid in tag_ids:
                            self.env.cr.execute(
                                "INSERT INTO sale_order_tag_rel "
                                "(order_id, tag_id) VALUES (%s, %s) "
                                "ON CONFLICT DO NOTHING",
                                (so.id, tid),
                            )
                        tagged += 1
                self.env.cr.execute(f"RELEASE SAVEPOINT {sp_name}")
            except Exception as e:  # noqa: BLE001
                # Rollback this SO's savepoint, keep the outer txn alive
                self.env.cr.execute(f"ROLLBACK TO SAVEPOINT {sp_name}")
                skipped_concurrent += 1
                _logger.warning(
                    "[KOB final stamp] skipped SO %s (%s): %s",
                    sn, so.id, e,
                )
        _logger.warning(
            "[KOB final stamp] %d orders updated, %d tagged, %d skipped "
            "(out of %d parsed)",
            updated, tagged, skipped_concurrent, len(by_sn),
        )

    def _normalize_date(self, value):
        if not value:
            return fields.Datetime.now()
        if isinstance(value, datetime):
            return fields.Datetime.to_string(
                value.replace(tzinfo=None) if value.tzinfo else value
            )
        s = str(value).strip().replace("/", "-")
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y %H:%M:%S",
                    "%d-%m-%Y"):
            try:
                return fields.Datetime.to_string(datetime.strptime(s, fmt))
            except ValueError:
                continue
        return fields.Datetime.now()

    # ── Main ──────────────────────────────────────────────────────
    def _collect_files(self):
        """Yield (filename, raw bytes, platform) tuples for every input
        file — single Excel/CSV via ``file_data`` AND the multi-file
        batch via ``attachment_ids``.  Platform is auto-detected from
        the filename; if it can't be detected we fall back to whatever
        is selected on the wizard."""
        if self.file_data:
            yield (
                self.filename or "upload.xlsx",
                base64.b64decode(self.file_data),
                self._detect_platform(self.filename) or self.platform,
            )
        for att in self.attachment_ids:
            raw = att.raw if att.raw else (
                base64.b64decode(att.datas) if att.datas else b""
            )
            if not raw:
                continue
            yield (
                att.name or "attachment.xlsx",
                raw,
                self._detect_platform(att.name) or self.platform,
            )

    def _notify_progress(self, done, total, last_label):
        """Update the wizard's progress_pct field for the form-bound
        progress bar widget. No more sticky toasts — the user sees
        progress on the wizard form itself, not as rapid-fire popups.
        """
        try:
            pct = int(done * 100 / total) if total else 100
            self.with_context(prefetch_fields=False).write({
                "progress_pct": pct,
            })
            # Commit so AJAX form refresh sees the new percentage.
            self.env.cr.commit()
        except Exception as e:
            _logger.debug("progress write skipped: %s", e)

    def action_validate(self):
        """Dry-run: parse files + check routing + duplicates + missing SKUs.
        Set state='preview' so user can review before confirming."""
        self.ensure_one()
        # Auto-fill fallbacks — these are only used when xlsx doesn't
        # carry the info per row.
        if not self.company_id:
            self.company_id = self.env.company
        if not self.warehouse_id:
            self.warehouse_id = self.env["stock.warehouse"].search([
                ("company_id", "=", self.company_id.id),
            ], limit=1)
        if not self.file_data and not self.attachment_ids:
            raise UserError(_(
                "Please upload at least one Excel/CSV file."
            ))
        records = []
        per_file_log = []
        for filename, raw, platform in self._collect_files():
            try:
                file_records = self._parse_blob(raw, filename)
            except UserError as e:
                per_file_log.append("FAIL %s — %s" % (filename, e.args[0]))
                continue
            for r in file_records:
                # Prefer per-row detection (Shopee_X / Lazada_X / Tiktok_X
                # in the Source/Shop column) so a single mixed-platform
                # xlsx imports correctly.
                r["__platform"] = r.get("_row_platform") or platform
                r["__source_file"] = filename
            per_file_log.append(
                "READ %s — %d rows (%s)"
                % (filename, len(file_records), platform),
            )
            records.extend(file_records)
        if not records:
            raise UserError(_(
                "No order rows could be parsed.\n\n%s"
            ) % "\n".join(per_file_log))
        by_order = {}
        for r in records:
            key = (r["__platform"], r["order_sn"])
            by_order.setdefault(key, []).append(r)

        ShopMap = self.env.get("wms.shop.company.map")
        Product = self.env["product.product"]
        all_companies = self.env["res.company"].sudo().search([
            ("active", "=", True),
        ])
        SaleOrder = self.env["sale.order"].with_context(
            allowed_company_ids=all_companies.ids,
        )

        def _resolve_company(platform, shop_name):
            if not ShopMap:
                return self.company_id
            pretty = _PRETTY.get(platform, platform.title())
            full = shop_name if (shop_name or "").lower().startswith(
                "%s_" % pretty.lower()
            ) else "%s_%s" % (pretty, shop_name or "")
            try:
                co = ShopMap.sudo().resolve(full)
            except Exception:
                co = None
            return co or self.company_id

        co_count = {}
        wh_count = {}
        unmapped = set()
        missing_skus = set()
        dup_count = 0
        Warehouse = self.env["stock.warehouse"].sudo()
        for (platform, order_sn), lines in by_order.items():
            shop = lines[0].get("shop") or ""
            target_co = _resolve_company(platform, shop)
            if ShopMap:
                pretty = _PRETTY.get(platform, platform.title())
                full = shop if (shop or "").lower().startswith(
                    "%s_" % pretty.lower()
                ) else "%s_%s" % (pretty, shop or "")
                if not ShopMap.sudo().search_count([("name", "=", full)]):
                    unmapped.add(full)

            # Per-row Warehouse override: if xlsx carries a Warehouse
            # column, search by exact name. Falls back to mapping when
            # not found.
            wh = Warehouse.browse()
            row_wh_name = lines[0].get("_row_warehouse")
            if row_wh_name:
                explicit = Warehouse.search(
                    [("name", "=", row_wh_name)], limit=1,
                )
                if explicit:
                    wh = explicit
                    target_co = explicit.company_id
            if not wh:
                wh = Warehouse.search([
                    ("company_id", "=", target_co.id),
                ], limit=1) or self.warehouse_id
                if (self.fbs_auto_route and platform == "shopee"
                        and order_sn.upper().endswith("-FBS")):
                    shopee_wh = Warehouse.search([
                        ("company_id", "=", target_co.id),
                        ("name", "ilike", "SHOPEE"),
                    ], limit=1)
                    if shopee_wh:
                        wh = shopee_wh
            co_count[target_co.name] = co_count.get(target_co.name, 0) + 1
            wh_count[wh.name] = wh_count.get(wh.name, 0) + 1
            if SaleOrder.sudo().search_count([
                "|",
                ("client_order_ref", "=", order_sn),
                ("name", "=", order_sn),
            ]):
                dup_count += 1
            for ln in lines:
                sku = ln.get("sku")
                if not sku:
                    continue
                if not Product.sudo().search_count([
                    "|", "|",
                    ("default_code", "=", sku),
                    ("x_kob_sku_code", "=", sku),
                    ("name", "=ilike", "[%s]%%" % sku),
                ]):
                    # Stash full Excel name + price next to the SKU so
                    # the preview can show exactly which product master
                    # row WOULD be created.
                    missing_skus.add((
                        sku,
                        (ln.get("raw_name") or "").strip()[:120],
                        float(ln.get("price") or 0),
                    ))

        co_lines = "\n".join(
            "  %s — %d orders" % (k, v) for k, v in sorted(co_count.items())
        ) or "  (none)"
        wh_lines = "\n".join(
            "  %s — %d orders" % (k, v) for k, v in sorted(wh_count.items())
        ) or "  (none)"
        # Render missing_skus as a review table — SKU · name · price.
        missing_text = ""
        if missing_skus:
            missing_text = (
                "⚠ The following SKUs are NOT in product master.\n"
                "   Tick \"I confirm auto-create\" below before Import "
                "to add them with the names + prices below.\n\n"
            )
            for sku, raw_name, price in sorted(missing_skus):
                missing_text += f"  [{sku}]  {raw_name}  →  {price:.2f} ฿\n"
        unmapped_text = ("\n".join(sorted(unmapped))
                         if unmapped else "")

        # ── Anomaly scan — surface every weird row BEFORE import ─────
        anomalies = self._kob_preview_anomalies(
            by_order, records, _resolve_company, Warehouse,
        )

        # Can proceed if either (a) every SKU exists in master, OR
        # (b) auto_create_product is on AND user has explicitly ticked
        # confirm_create_missing after reviewing the list above.
        # Also requires no BLOCKING anomalies.
        blocking = any(a["severity"] == "BLOCK" for a in anomalies)
        can_proceed = bool(by_order) and (
            not missing_skus or (
                self.auto_create_product and self.confirm_create_missing
            )
        ) and not blocking

        # Build human-readable anomaly section for preview_log.
        anom_lines = []
        if anomalies:
            anom_lines.append("")
            anom_lines.append("═" * 60)
            anom_lines.append("⚠️  ANOMALY REPORT — review before clicking Import")
            anom_lines.append("═" * 60)
            by_sev = {}
            for a in anomalies:
                by_sev.setdefault(a["severity"], []).append(a)
            for sev in ("BLOCK", "WARN", "INFO"):
                if sev not in by_sev:
                    continue
                icon = {"BLOCK": "⛔", "WARN": "⚠️", "INFO": "ℹ️"}[sev]
                anom_lines.append("")
                anom_lines.append(f"{icon} {sev}  ({len(by_sev[sev])} item(s))")
                for a in by_sev[sev][:30]:  # cap per-severity at 30 rows
                    anom_lines.append(f"   • {a['msg']}")
                if len(by_sev[sev]) > 30:
                    anom_lines.append(
                        f"   … and {len(by_sev[sev]) - 30} more"
                    )

        self.write({
            "state": "preview",
            "preview_orders_count": len(by_order),
            "preview_lines_count": len(records),
            "preview_duplicates_count": dup_count,
            "preview_missing_skus": missing_text,
            "preview_unmapped_shops": unmapped_text,
            "preview_companies_breakdown": co_lines,
            "preview_warehouses_breakdown": wh_lines,
            "preview_can_proceed": can_proceed,
            "preview_log": "\n".join(per_file_log + anom_lines),
        })
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def action_back_to_edit(self):
        self.ensure_one()
        self.write({
            "state": "draft",
            "preview_orders_count": 0,
            "preview_lines_count": 0,
            "preview_duplicates_count": 0,
            "preview_missing_skus": False,
            "preview_unmapped_shops": False,
            "preview_companies_breakdown": False,
            "preview_warehouses_breakdown": False,
            "preview_can_proceed": False,
            "preview_log": False,
        })
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    def action_import(self):
        self.ensure_one()
        if self.state != "preview":
            raise UserError(_(
                "กด Validate ก่อน. ต้อง preview ดูสรุปก่อน Confirm Import."
            ))
        if not self.preview_can_proceed:
            raise UserError(_(
                "❌ Preview พบปัญหาต้องแก้ก่อน import:\n\n"
                "• มี SKU ที่ไม่อยู่ใน product master\n"
                "  → ดูรายการ ⚠ Missing SKUs ด้านบน\n"
                "  → ติ๊ก '✔ I confirm auto-create' เพื่อยอมรับชื่อ + ราคา\n"
                "    จาก Excel เป็น product master record ใหม่\n"
                "  → หรือ เพิ่ม product ใน master ก่อน แล้วกด Back → "
                "Validate ใหม่\n\n"
                "• อาจมี BLOCK anomaly ใน Anomaly Report — ดูที่ "
                "Preview Log แก้แล้ว validate ใหม่"
            ))
        # ── Explicit company selection guard ────────────────────────
        # Without this, env.company silently routed every BTV file into
        # KOB. A required field on the model isn't enough — multi-company
        # users may still trigger the wizard with the wrong env.company.
        if not self.company_id:
            raise UserError(_(
                "⚠ ต้องเลือกบริษัทก่อน import\n\n"
                "Please tick a Company before importing. The marketplace "
                "import will create Sales Orders + Pickings under the "
                "selected company. Choosing the wrong company moves all "
                "stock movement and invoicing to the wrong books."
            ))
        if not self.warehouse_id:
            raise UserError(_(
                "⚠ ต้องเลือก Warehouse ของบริษัท %s ก่อน import"
            ) % self.company_id.name)
        if self.warehouse_id.company_id != self.company_id:
            raise UserError(_(
                "⚠ Warehouse %s อยู่ในบริษัท %s แต่คุณเลือกบริษัท %s — "
                "ไม่ตรงกัน เลือก warehouse ใหม่ให้ตรงกับบริษัท"
            ) % (
                self.warehouse_id.name,
                self.warehouse_id.company_id.name,
                self.company_id.name,
            ))
        # Allow the direct-records path (action_import_records) to bypass
        # the file-required guard.
        if (not self.file_data
                and not self.attachment_ids
                and not self.env.context.get("_rds_records")):
            raise UserError(_(
                "Please upload at least one Excel/CSV file (single or "
                "drop the whole folder into 'Excel/CSV Files (batch)')."
            ))

        # ── Records source: pre-shaped (action_import_records) or files ─
        # `_rds_records` is set in the env context by
        # `action_import_records()` so callers can skip the xlsx parse
        # entirely and ship already-shaped dicts over XML-RPC. When the
        # key is absent we go through the legacy parse loop.
        records = self.env.context.get("_rds_records")
        per_file_log = self.env.context.get("_rds_per_file_log")
        if records is None:
            # Parse every input file.  Each row remembers its platform so
            # we use the right partner / source / warehouse.
            records = []
            per_file_log = []
            for filename, raw, platform in self._collect_files():
                try:
                    file_records = self._parse_blob(raw, filename)
                except UserError as e:
                    per_file_log.append("FAIL %s — %s" % (filename, e.args[0]))
                    continue
                for r in file_records:
                    # Prefer per-row detection (Shopee_X / Lazada_X / Tiktok_X
                    # in the Source/Shop column) so a single mixed-platform
                    # xlsx imports correctly.
                    r["__platform"] = r.get("_row_platform") or platform
                    r["__source_file"] = filename
                per_file_log.append(
                    "READ %s — %d rows (%s)"
                    % (filename, len(file_records), platform),
                )
                records.extend(file_records)

        if not records:
            raise UserError(_(
                "No order rows could be parsed.\n\n%s"
            ) % "\n".join(per_file_log or []))

        # Group rows by (platform, order_sn) — a single order_sn could
        # legally appear under different platforms.
        by_order = {}
        for r in records:
            key = (r["__platform"], r["order_sn"])
            by_order.setdefault(key, []).append(r)

        fake_tag = self.env.ref("kob_marketplace_import.tag_fake_order")
        # Beauty Vill product tag — products bearing it route to BTV-WH2 (Online).
        btv_tag = self.env.ref(
            "kob_marketplace_import.product_tag_beauty_vill",
            raise_if_not_found=False,
        )
        btv_wh = self.env["stock.warehouse"].search(
            [("code", "=", "B-On")], limit=1,
        ) if btv_tag else self.env["stock.warehouse"]
        SaleOrder = self.env["sale.order"].with_company(self.company_id)
        log_lines = list(per_file_log)
        sales = self.env["sale.order"]

        # ── Progress tracking ─────────────────────────────────────────
        total = len(by_order)
        self.write({
            "state": "running",
            "progress_total": total,
            "progress_done": 0,
            "progress_pct": 0,
            "imported_count": 0,
            "duplicate_count": 0,
            "unresolved_count": 0,
        })
        self.env.cr.commit()
        self._notify_progress(0, total, "Starting…")
        completed = 0
        imported_n = 0
        duplicate_n = 0
        unresolved_n = 0
        duplicate_so_ids = []
        duplicate_lines = []  # detailed log for duplicate_summary field

        # Cache partner per platform to avoid repeated searches.
        partner_cache = {}

        def _partner_for(platform):
            if platform in partner_cache:
                return partner_cache[platform]
            name = _PARTNER_BY_PLATFORM[platform]
            partner = self.env["res.partner"].search([("name", "=", name)],
                                                     limit=1)
            if not partner:
                partner = self.env["res.partner"].create({
                    "name": name, "is_company": True, "customer_rank": 1,
                })
            partner_cache[platform] = partner
            return partner

        def _source_for(platform, shop_name):
            pretty = _PRETTY[platform]
            # Employee orders use a single "Employee" utm.source — the
            # Excel `shop` column for staff perks is usually empty or a
            # placeholder (Lazada_Unknown) that we don't want trailing
            # into the source name.
            if platform == "employee":
                full = "Employee"
            elif shop_name.lower().startswith("%s_" % pretty.lower()):
                full = shop_name
            else:
                full = "%s_%s" % (pretty, shop_name)
            s = self.env["utm.source"].search([("name", "=", full)], limit=1)
            if not s:
                s = self.env["utm.source"].create({"name": full})
            return s

        for (platform, order_sn), lines in by_order.items():
            # Robust duplicate check: match either client_order_ref OR
            # name (marketplace import sets SO.name = order_sn too).
            existing = SaleOrder.search([
                "|",
                ("client_order_ref", "=", order_sn),
                ("name", "=", order_sn),
            ], limit=1)
            if existing:
                duplicate_n += 1
                duplicate_so_ids.append(existing.id)
                detail = (
                    f"{order_sn}  →  existing SO {existing.name} "
                    f"(state={existing.state}, "
                    f"date={existing.date_order and existing.date_order.strftime('%Y-%m-%d') or '—'}, "
                    f"customer={existing.partner_id.name or '—'})"
                )
                duplicate_lines.append(detail)
                log_lines.append(f"SKIP {order_sn} — already imported "
                                 f"({existing.name})")
                # Still count toward progress so the bar reaches 100%.
                completed += 1
                pct = int(completed * 100 / total) if total else 100
                self.write({
                    "progress_done": completed,
                    "progress_pct": pct,
                    "duplicate_count": duplicate_n,
                })
                if completed % 5 == 0 or completed == total:
                    self.env.cr.commit()
                self._notify_progress(
                    completed, total,
                    f"Skipped duplicate {order_sn}",
                )
                continue

            partner = _partner_for(platform)
            head = lines[0]
            shop = head["shop"] or "Unknown"
            source = _source_for(platform, shop)
            # Collect raw tag labels from every row's "Tags" column for
            # later resolve+write AFTER SO create (see post_write block).
            # Brackets are preserved verbatim so "Online[FS]" survives.
            import re as _re
            tag_label_sources = []
            for ln in lines:
                if ln.get("tags"):
                    tag_label_sources.append(ln["tags"])

            order_lines_vals = []
            stock_adjust_queue = []  # (product, qty, unit_cost) for after target_wh
            for ln in lines:
                product = self._resolve_product(
                    ln["sku"], ln.get("brand"),
                    excel_name=ln.get("raw_name"),
                    excel_price=ln.get("price"),
                )
                if not product:
                    log_lines.append(
                        f"WARN {order_sn} — SKU '{ln['sku']}' not found, "
                        f"line skipped (toggle 'Auto-create missing "
                        f"products' to import it anyway)",
                    )
                    continue
                # Defer stock adjustment until target_wh is resolved
                # below (FBS / BTV routing can swap warehouse).
                # Service / non-storable products (fees, shipping, platform
                # surcharges) carry no stock — skip the queue entry.
                _is_storable = True
                if "is_storable" in product._fields:
                    _is_storable = bool(product.is_storable)
                _is_service = (product.type == "service") or (not _is_storable)
                if self.auto_adjust_lot and not _is_service:
                    stock_adjust_queue.append((
                        product,
                        float(ln.get("qty") or 0),
                        float(ln.get("price") or 0),
                    ))
                # Price: keep marketplace precision verbatim. "Product Price"
                # decimal_precision is forced to 4 digits via data file so the
                # ORM stores full sub-satang accuracy (e.g. Shopee 36.292835
                # is preserved without rounding compound loss across lines).
                # `price` of 0 still marks a free gift.
                excel_price = float(ln.get("price") or 0)
                price_unit = excel_price
                vals = {
                    "product_id": product.id,
                    "product_uom_qty": ln["qty"],
                    "price_unit": price_unit,
                }
                if excel_price == 0:
                    vals["name"] = f"{product.display_name}  🎁 Free Gift"
                # VAT 7%: enforce on every imported line for the SO's
                # routed company (KOB/BTV/CMN each have their own tax id).
                vat_tax = self.env["account.tax"].sudo().search([
                    ("company_id", "=", self.company_id.id),
                    ("amount", "=", 7),
                    ("type_tax_use", "=", "sale"),
                    ("name", "ilike", "Output VAT"),
                ], limit=1)
                if vat_tax:
                    line_fields = self.env["sale.order.line"]._fields
                    tax_field = "tax_ids" if "tax_ids" in line_fields else "tax_id"
                    vals[tax_field] = [(6, 0, [vat_tax.id])]
                # UoM from Excel (Order Lines/Unit of Measure column).
                # Default to the product's UoM when blank or unmatched.
                uom_label = ln.get("uom")
                if uom_label:
                    uom = self.env["uom.uom"].search(
                        [("name", "=ilike", uom_label)], limit=1,
                    )
                    if uom:
                        # Odoo 19 renamed sale.order.line.product_uom →
                        # product_uom_id; fall back if older API.
                        line_model = self.env["sale.order.line"]
                        if "product_uom_id" in line_model._fields:
                            vals["product_uom_id"] = uom.id
                        elif "product_uom" in line_model._fields:
                            vals["product_uom"] = uom.id
                order_lines_vals.append((0, 0, vals))

            if not order_lines_vals:
                unresolved_n += 1
                log_lines.append(f"SKIP {order_sn} — all lines unresolved")
                completed += 1
                pct = int(completed * 100 / total) if total else 100
                self.write({
                    "progress_done": completed,
                    "progress_pct": pct,
                    "unresolved_count": unresolved_n,
                })
                if completed % 5 == 0 or completed == total:
                    self.env.cr.commit()
                self._notify_progress(
                    completed, total,
                    f"Skipped (unresolved) {order_sn}",
                )
                continue

            # FBS auto-routing: if order_sn ends with "-FBS", route to the
            # company's *-SHOPEE warehouse (Shopee Fulfilled-By-Shopee).
            target_wh = self.warehouse_id
            if (self.fbs_auto_route
                and platform == "shopee"
                and order_sn.upper().endswith("-FBS")):
                shopee_wh = self.env["stock.warehouse"].search([
                    ("company_id", "=", self.company_id.id),
                    ("name", "ilike", "SHOPEE"),
                ], limit=1)
                if shopee_wh:
                    target_wh = shopee_wh

            # Beauty Vill tag → BTV-WH2 (Online) + BTV company. Wins over
            # warehouse_id but yields to FBS routing above. Swapping the
            # company keeps SO.company_id == warehouse.company_id so create
            # passes the cross-company validation.
            so_company_id = self.company_id.id
            if (btv_tag and btv_wh
                and not (platform == "shopee"
                         and order_sn.upper().endswith("-FBS"))):
                line_products = self.env["product.product"].browse(
                    [v[2]["product_id"] for v in order_lines_vals
                     if v and v[2] and v[2].get("product_id")]
                )
                if any(btv_tag.id in p.product_tag_ids.ids
                       or btv_tag.id in p.product_tmpl_id.product_tag_ids.ids
                       for p in line_products):
                    target_wh = btv_wh
                    so_company_id = btv_wh.company_id.id

            # ── Stock adjustment NOW (after target_wh known) ─────────────
            # Top up lot + quant at the RESOLVED warehouse so SO confirm
            # can reserve.  Re-running on an already-stocked lot is safe
            # (target_qty accumulates per import; duplicate detection
            # blocks re-import of the same order_sn).
            for product, qty, unit_cost in stock_adjust_queue:
                try:
                    self._ensure_stock_with_lot(
                        product, qty, target_wh,
                        unit_cost=unit_cost, order_sn=order_sn,
                    )
                except Exception as e:
                    _logger.exception(
                        "Stock adjust failed for %s qty=%s wh=%s: %s",
                        product.default_code or product.id,
                        qty, target_wh.code, e,
                    )

            # Use the platform order number as the SO reference so the
            # operations team scans / searches the same identifier
            # everywhere (Excel, label, picking, accounting).  Odoo
            # accepts an explicit ``name`` value at create time and
            # skips the ir.sequence generation.  We still keep
            # ``client_order_ref`` populated for backwards-compatible
            # reporting.
            so_name = str(order_sn).strip() or False
            so_vals = {
                "name":             so_name,
                "partner_id":       partner.id,
                "client_order_ref": order_sn,
                "company_id":       so_company_id,
                "warehouse_id":     target_wh.id,
                "date_order":       self._normalize_date(head["order_date"]),
                "source_id":        source.id,
                "order_line":       order_lines_vals,
            }
            # Delivery Date (Excel column "Delivery Date") → commitment_date.
            if head.get("delivery_date"):
                so_vals["commitment_date"] = self._normalize_date(
                    head["delivery_date"]
                )
            # Source Document (Excel) → sale.order.origin so downstream
            # pickings / invoices carry the marketplace reference.
            if head.get("source_document"):
                so_vals["origin"] = str(head["source_document"]).strip()
            # Order Type from Excel — lookup by name; fall back to Normal Order.
            order_type = None
            ot_label = head.get("order_type")
            if ot_label:
                order_type = self.env["sale.order.type"].search(
                    [("name", "=ilike", ot_label)], limit=1,
                )
            if not order_type:
                order_type = self.env.ref(
                    "kob_marketplace_import.sale_order_type_normal",
                    raise_if_not_found=False,
                )
            if order_type:
                so_vals["sale_order_type_id"] = order_type.id
            # Sales Team — auto-resolve to the right team for the
            # routed company. Look up by Excel label (or "eMarketplace"
            # default) scoped to so_company_id; if no match exists in
            # this company, **auto-create** one so cross-company SOs
            # always have a valid team without manual setup.
            st_label = (head.get("sales_team") or "eMarketplace").strip()
            CrmTeam = self.env["crm.team"]
            team = CrmTeam.search([
                "|", ("company_id", "=", so_company_id),
                     ("company_id", "=", False),
                ("name", "=ilike", st_label),
            ], limit=1)
            if not team:
                # Auto-create the team inside the routed company so the
                # imported SO has a Sales Team that won't trigger the
                # cross-company validator.
                team = CrmTeam.sudo().create({
                    "name": st_label,
                    "company_id": so_company_id,
                })
            so_vals["team_id"] = team.id
            # Re-bind company context if the order was BTV-routed.
            so_creator = (
                self.env["sale.order"].with_company(so_company_id)
                if so_company_id != self.company_id.id
                else SaleOrder
            )
            so = so_creator.create(so_vals)
            _logger.warning(
                "[KOB DEBUG] SO %s head.keys=%s tags=%r delivery=%r src_doc=%r order_date=%r",
                order_sn, list(head.keys()),
                head.get("tags"), head.get("delivery_date"),
                head.get("source_document"), head.get("order_date"),
            )

            # ── Rounding reconciliation ─────────────────────────────────
            # When the Excel ships a marketplace "Order Total" (customer-paid)
            # column, nudge the last line's price_unit so Odoo's computed
            # amount_total matches it exactly. Marketplaces export
            # already-rounded 2-decimal prices, then their VAT math gives a
            # gross total that Odoo can never reproduce by recomputing from
            # line prices (e.g. 3 × 127.41 × 1.07 = 408.99 here, but the
            # customer was charged 409.00). Absorbing the gap (≤ a few satang)
            # into the last line keeps tax computation honest while making
            # bank-reconciliation 1:1 with the marketplace report.
            expected_total = head.get("order_total")
            if expected_total and so.order_line:
                # Currency precision for THB = 2 decimals; gap below 0.005
                # is rounding noise we can ignore.
                gap = float(expected_total) - so.amount_total
                if abs(gap) >= 0.005 and abs(gap) < 1.0:
                    # Apply the correction to the last non-zero-priced line so
                    # free-gift lines (price=0) keep their semantic.
                    paid_lines = so.order_line.filtered(
                        lambda l: l.price_unit and l.product_uom_qty
                    )
                    target = paid_lines[-1] if paid_lines else so.order_line[-1]
                    # Solve for the per-unit delta that, after VAT, closes the
                    # gap. With VAT 7% inclusive of the recomputation, the
                    # untaxed delta = gap / 1.07; per-unit = delta / qty.
                    tax_factor = 1.0
                    if target.tax_ids:
                        tax_factor = 1 + sum(t.amount or 0 for t in target.tax_ids) / 100.0
                    untaxed_delta = gap / tax_factor
                    per_unit_delta = untaxed_delta / (target.product_uom_qty or 1.0)
                    new_price = target.price_unit + per_unit_delta
                    target.with_company(so_company_id).write({
                        "price_unit": new_price,
                    })
                    _logger.info(
                        "[KOB] Rounding nudge %s: gap=%.4f → line %s price_unit %s → %s",
                        order_sn, gap, target.id, target.price_unit - per_unit_delta, new_price,
                    )
            # Force commitment_date + origin + tag_ids AFTER create —
            # sale.order's compute methods (_compute_commitment_date,
            # tag_ids onchange) can wipe values passed via vals.
            post_write = {}
            if head.get("delivery_date"):
                post_write["commitment_date"] = self._normalize_date(
                    head["delivery_date"]
                )
            if head.get("source_document"):
                post_write["origin"] = str(head["source_document"]).strip()
            if head.get("order_date"):
                post_write["date_order"] = self._normalize_date(
                    head["order_date"]
                )
            # Tag write — explicit (6, 0, [...]) replaces the m2m entirely
            # so the result matches the Excel "Tags" column 1:1.
            tag_id_list = []
            if any(l["fake"] for l in lines):
                tag_id_list.append(fake_tag.id)
            for raw in tag_label_sources:
                for label in [t.strip() for t in _re.split(r"[,;|]", raw) if t.strip()]:
                    tag = self.env["crm.tag"].search(
                        [("name", "=ilike", label)], limit=1,
                    )
                    if not tag:
                        tag = self.env["crm.tag"].sudo().create({"name": label})
                    if tag.id not in tag_id_list:
                        tag_id_list.append(tag.id)
            if tag_id_list:
                post_write["tag_ids"] = [(6, 0, tag_id_list)]
            if post_write:
                try:
                    so.with_company(so_company_id).write(post_write)
                except Exception as e:
                    _logger.exception(
                        "post_write failed for %s: %s", so.name, e,
                    )
            sales |= so
            log_lines.append(f"OK   {order_sn} → {so.name} ({len(so.order_line)} lines)")

            # Honor the upstream RDS state when present on the record:
            #   _rds_state='draft' → keep kob SO as Quotation
            #   _rds_state='cancel' → cancel the kob SO too
            #   _rds_state='sale' (or absent) → auto_confirm as before
            rds_state = head.get("_rds_state") or "sale"
            if self.auto_confirm and rds_state == "sale" and so.state in ("draft", "sent"):
                # Guard: action_confirm() raises "Some orders are not in a
                # state requiring confirmation" if the SO is already in
                # 'sale' / 'cancel' state.  In bulk imports a single such
                # row otherwise poisons the whole wizard transaction and
                # every prior SO in the same batch gets rolled back.
                try:
                    so.action_confirm()
                except UserError as _confirm_err:
                    _logger.warning(
                        "auto_confirm skipped for %s — %s",
                        so.name, _confirm_err.args[0] if _confirm_err.args else _confirm_err,
                    )
            elif rds_state == "cancel" and so.state != "cancel":
                # Mirror source cancellation.
                try:
                    so._action_cancel()
                except Exception as _cancel_err:
                    _logger.warning(
                        "RDS-driven cancel failed for %s — %s",
                        so.name, _cancel_err,
                    )
                # Re-stamp commitment_date / origin / tag_ids AFTER
                # action_confirm because sale.order's compute chain can
                # overwrite the values we set above. SQL UPDATE skips
                # ORM hooks so the values stick.
                if post_write:
                    try:
                        sql_sets = []
                        sql_args = []
                        if "commitment_date" in post_write:
                            sql_sets.append("commitment_date=%s")
                            sql_args.append(post_write["commitment_date"])
                        if "origin" in post_write:
                            sql_sets.append("origin=%s")
                            sql_args.append(post_write["origin"])
                        if "date_order" in post_write:
                            sql_sets.append("date_order=%s")
                            sql_args.append(post_write["date_order"])
                        if sql_sets:
                            sql_args.append(so.id)
                            self.env.cr.execute(
                                "UPDATE sale_order SET " + ",".join(sql_sets)
                                + " WHERE id=%s",
                                sql_args,
                            )
                        if tag_id_list:
                            # Rebuild the m2m link rows directly.
                            self.env.cr.execute(
                                "DELETE FROM sale_order_tag_rel WHERE order_id=%s",
                                (so.id,),
                            )
                            for tid in tag_id_list:
                                self.env.cr.execute(
                                    "INSERT INTO sale_order_tag_rel "
                                    "(order_id, tag_id) VALUES (%s, %s) "
                                    "ON CONFLICT DO NOTHING",
                                    (so.id, tid),
                                )
                    except Exception as e:
                        _logger.exception(
                            "SQL re-stamp failed for %s: %s", so.name, e,
                        )
                if self.auto_assign:
                    for picking in so.picking_ids:
                        picking.action_assign()
                        # Stamp x_kob_* fields onto the picking.
                        picking.write({
                            "x_kob_source_ref":     source.id,
                            "x_kob_order_date_ref": so.date_order,
                            "x_kob_fake_order":     any(l["fake"] for l in lines),
                        })
                        # Propagate brand product → move.
                        for move in picking.move_ids:
                            brand = (move.sale_line_id and
                                     move.sale_line_id.x_kob_brand) \
                                    or move.product_id.x_kob_brand
                            if brand:
                                move.write({"x_kob_brand": brand})
                        # Bridge: create the wms.sales.order record so the
                        # KOB WMS Orders list / Pick / Pack screens see it.
                        if hasattr(picking, "action_create_wms_order"):
                            try:
                                picking.action_create_wms_order()
                            except Exception as e:
                                _logger.warning(
                                    "WMS bridge failed for %s: %s",
                                    picking.name, e,
                                )

            # ── Per-order progress: update + commit + notify ─────────
            completed += 1
            imported_n += 1
            pct = int(completed * 100 / total) if total else 100
            self.write({
                "progress_done": completed,
                "progress_pct": pct,
                "imported_count": imported_n,
            })
            # Commit every 5 orders (or last) so a refresh shows progress
            # and a crash doesn't lose imported orders. Notify on every
            # order so the user sees a live toast counter.
            if completed % 5 == 0 or completed == total:
                self.env.cr.commit()
            self._notify_progress(completed, total,
                                   f"Imported {order_sn}")

        # ── FINAL BACKFILL — guaranteed Excel→SO field stamping ──────
        # Independent of per-order create/confirm flow: read every parsed
        # record one more time and force-set commitment_date / origin /
        # date_order / tag_ids via raw SQL. Bypasses ORM compute methods
        # and any rollback that may have wiped the in-loop post_write.
        try:
            self._kob_final_stamp(records)
            self.env.cr.commit()
        except Exception as e:
            _logger.exception("Final stamp failed: %s", e)

        # ── Final summary write ───────────────────────────────────────
        dup_summary = (
            "\n".join(duplicate_lines) if duplicate_lines
            else _("No duplicates in this batch — every order_sn was new.")
        )
        self.write({
            "state":            "done",
            "log":              "\n".join(log_lines),
            "sale_order_ids":   [(6, 0, sales.ids)],
            "duplicate_order_ids": [(6, 0, list(set(duplicate_so_ids)))],
            "duplicate_summary":   dup_summary,
            "progress_pct":     100,
            "imported_count":   imported_n,
            "duplicate_count":  duplicate_n,
            "unresolved_count": unresolved_n,
        })
        self.env.cr.commit()

        # Snapshot results into a persistent session record so the user
        # has history + drill-down + auto-batch link, instead of a
        # transient wizard that vacuums.
        session = self.env["kob.marketplace.import.session"].sudo().create({
            "user_id": self.env.user.id,
            "platform": self.platform,
            "company_id": self.company_id.id,
            "warehouse_id": self.warehouse_id.id if self.warehouse_id else False,
            "date_started": getattr(self, "_kob_t0", False) or fields.Datetime.now(),
            "date_finished": fields.Datetime.now(),
            "file_count": 1 + len(self.batch_attachment_ids) if hasattr(self, "batch_attachment_ids") else 1,
            "filenames": (self.filename or "") + ("\n" + "\n".join(
                a.name for a in self.batch_attachment_ids) if hasattr(self, "batch_attachment_ids") and self.batch_attachment_ids else ""),
            "total_rows": total,
            "orders_attempted": total,
            "orders_created": imported_n,
            "orders_skipped": duplicate_n,
            "orders_failed": unresolved_n,
            "sale_order_ids": [(6, 0, sales.ids)],
            "log": "\n".join(log_lines),
            "state": "done",
        })

        # Auto-create WMS courier batch grouping the freshly-imported SOs
        # by courier so the dispatch worker can ship the whole import
        # round in one click instead of opening every order.
        try:
            confirmed_sos = sales.filtered(
                lambda s: s.state in ("sale", "done") and s.picking_ids)
            couriers = {}
            for so in confirmed_sos:
                for pick in so.picking_ids.filtered(
                        lambda p: p.state in ("assigned", "done")):
                    if pick.kob_carrier_id:
                        couriers.setdefault(pick.kob_carrier_id.id, set()).add(so.id)
            if confirmed_sos and "wms.courier.batch" in self.env.registry.models:
                Batch = self.env["wms.courier.batch"].sudo()
                lead_courier = next(iter(couriers), False) if couriers else False
                courier = self.env["wms.courier"].browse(lead_courier) \
                    if lead_courier else self.env["wms.courier"].search([], limit=1)
                if courier:
                    # Stamp round metadata so kob_wms.create() builds the
                    # ROUND-#/COURIER/DATE name instead of falling back to
                    # the generic BCH-XXXXX ir.sequence.
                    round_date, round_no = Batch._compute_round_date()
                    batch = Batch.create({
                        "courier_id": courier.id,
                        "state": "scanning",
                        "dispatch_round_date": round_date,
                        "dispatch_round_number": round_no,
                        "note": f"Auto-created from import session {session.name}",
                    })
                    session.write({"wms_batch_id": batch.id})
        except Exception as exc:  # noqa: BLE001
            _logger.info("Auto-batch skipped for session %s: %s", session.name, exc)

        return {
            "type": "ir.actions.act_window",
            "name": _("Import Session — %s") % session.name,
            "res_model": "kob.marketplace.import.session",
            "res_id": session.id,
            "view_mode": "form",
            "target": "current",
        }

    def action_import_records(self, records_in):
        """External-API entry: import a list of pre-shaped record dicts
        without going through the xlsx / csv parse stage.

        Expected per-record keys (same shape as ``_parse_blob`` output):
            order_sn (required), sku (required), qty (required),
            shop, order_date, delivery_date, order_type, customer,
            sales_team, tags, source_document, external_id,
            _row_platform, _row_warehouse, raw_name, uom, brand, fake,
            price, order_total.

        The wizard must already be configured with ``company_id`` +
        ``warehouse_id``.  State is forced to ``preview`` and
        ``preview_can_proceed`` is set to True so the standard
        ``action_import`` guard logic still runs but the file-upload
        check is bypassed.

        Returns whatever ``action_import`` returns (typically an
        ir.actions.act_window dict pointing at the created session).
        """
        self.ensure_one()
        if not records_in:
            raise UserError(_("action_import_records called with empty list."))
        # Annotate platform fallback + source label so the downstream
        # grouping logic in action_import sees the same shape it would
        # from _parse_blob + the post-parse augmentation.
        annotated = []
        for r in records_in:
            r = dict(r)  # don't mutate caller's data
            r.setdefault("__platform",
                         r.get("_row_platform") or self.platform)
            r.setdefault("__source_file", "rds_direct")
            annotated.append(r)
        # Bypass the "click Validate first" guard at the top of
        # action_import — the records have already been validated by
        # the upstream SQL extractor.
        self.write({
            "state": "preview",
            "preview_can_proceed": True,
        })
        # Stash records in env context (Odoo records don't allow ad-hoc
        # Python attributes) — action_import picks them up via
        # ``self.env.context.get`` in lieu of running the file parse loop.
        return self.with_context(
            _rds_records=annotated,
            _rds_per_file_log=["READ rds_direct — %d rows" % len(annotated)],
        ).action_import()

    def action_open_orders(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "name": _("Imported Sales Orders"),
            "res_model": "sale.order",
            "view_mode": "list,form",
            "domain": [("id", "in", self.sale_order_ids.ids)],
        }

    def action_open_duplicates(self):
        """Open the SOs that were skipped because they already existed."""
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "name": _("Skipped (Already Imported)"),
            "res_model": "sale.order",
            "view_mode": "list,form",
            "domain": [("id", "in", self.duplicate_order_ids.ids)],
        }
