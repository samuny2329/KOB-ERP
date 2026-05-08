"""Import Pick Groups wizard.

Reads the Excel produced by ``Print_Label-App 1.0.3`` (sheet exported by
``A - First Step - Extractions/Start-Delivery.py``) and turns each printed
pick set into a ``wms.pick.group`` record with member orders linked.

Expected columns (header row 1):
    index, name, brand_sort, sku_sort, x_kob_fake_order,
    x_kob_order_date_ref, x_kob_source_ref, running, result_B

Grouping logic:
    - ``running`` field has the form ``"No.{set_num}-{pos}/{total}"``
    - All rows with the same ``set_num`` belong to the same pick set
    - ``sku_sort`` is identical across members of a set
    - Member ``wms.sales.order`` is matched by ``ref`` ↔ ``x_kob_source_ref``
"""
import base64
import io
import logging
import re
from collections import defaultdict

from odoo import api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

REQUIRED_COLS = ("sku_sort", "x_kob_source_ref", "running")
RUNNING_RE = re.compile(r"No\.(\d+)-(\d+)/(\d+)")


class WmsPickGroupImportWizard(models.TransientModel):
    _name = "wms.pick.group.import.wizard"
    _description = "WMS Pick Group Import Wizard (Print_Label-App xlsx)"

    xlsx_file = fields.Binary(
        string="Print_Label-App Excel", required=True)
    xlsx_filename = fields.Char()
    overwrite = fields.Boolean(
        string="Overwrite Existing Group Assignment",
        default=False,
        help="When checked, reassigns SOs even if they already belong to "
             "another pick group. Default off — leave existing assignments "
             "alone.",
    )
    promote_orders = fields.Boolean(
        string="Promote member orders to 'picking'",
        default=False,
        help="If checked, member SOs currently in 'pending' state are "
             "advanced to 'picking' once linked to the group. Use sparingly.",
    )

    summary = fields.Text(readonly=True)

    def action_import(self):
        self.ensure_one()
        if not self.xlsx_file:
            raise UserError("Attach the Print_Label-App Excel first.")

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise UserError(
                "openpyxl is required. Install with: "
                "docker exec kob-odoo-19 pip install openpyxl"
            ) from exc

        raw = base64.b64decode(self.xlsx_file)
        try:
            wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise UserError(f"Could not read xlsx: {exc}") from exc

        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration as exc:
            raise UserError("Excel is empty.") from exc

        col_idx = {name: i for i, name in enumerate(header) if name}
        missing = [c for c in REQUIRED_COLS if c not in col_idx]
        if missing:
            raise UserError(
                f"Missing required columns: {missing}. "
                f"Got: {list(header)}"
            )

        # Group rows by set_num
        sets = defaultdict(list)  # set_num -> [{row data}]
        scanned = 0
        skipped_no_running = 0
        for row in rows_iter:
            scanned += 1
            running = row[col_idx["running"]]
            if not running or not isinstance(running, str):
                skipped_no_running += 1
                continue
            m = RUNNING_RE.match(running.strip())
            if not m:
                skipped_no_running += 1
                continue
            set_num = int(m.group(1))
            sets[set_num].append({
                "set_num": set_num,
                "pos": int(m.group(2)),
                "total": int(m.group(3)),
                "sku_sort": row[col_idx["sku_sort"]] or "",
                "source_ref": (
                    row[col_idx["x_kob_source_ref"]] or "").strip()
                if isinstance(row[col_idx["x_kob_source_ref"]], str)
                else (row[col_idx["x_kob_source_ref"]] or ""),
            })

        if not sets:
            raise UserError(
                f"No rows with valid 'running' (No.X-Y/Z) field found. "
                f"Scanned {scanned} rows."
            )

        Group = self.env["wms.pick.group"]
        SO = self.env["wms.sales.order"]
        groups_created = 0
        groups_existing = 0
        orders_assigned = 0
        orders_unmatched = []
        orders_skipped_existing = 0

        for set_num in sorted(sets):
            members = sets[set_num]
            sku_sort = members[0]["sku_sort"] or ""
            # Group key = sku_sort (stable across imports for same product set)
            key = sku_sort or f"set-{set_num}"
            grp = Group.search([("name", "=", key),
                                ("source", "=", "print_label_app")], limit=1)
            if not grp:
                grp = Group.create({
                    "name": key,
                    "sku_sort": sku_sort,
                    "source": "print_label_app",
                })
                groups_created += 1
            else:
                groups_existing += 1

            for m in members:
                ref = str(m["source_ref"]).strip()
                if not ref:
                    continue
                so = SO.search([("ref", "=", ref)], limit=1)
                if not so:
                    orders_unmatched.append(ref)
                    continue
                if so.pick_group_id and so.pick_group_id != grp:
                    if not self.overwrite:
                        orders_skipped_existing += 1
                        continue
                vals = {"pick_group_id": grp.id}
                if self.promote_orders and so.status == "pending":
                    vals["status"] = "picking"
                so.write(vals)
                orders_assigned += 1

        lines = [
            f"Rows scanned: {scanned}",
            f"Rows skipped (no valid 'running'): {skipped_no_running}",
            f"Sets in file: {len(sets)}",
            f"Groups created: {groups_created}",
            f"Groups existing (reused): {groups_existing}",
            f"Orders assigned: {orders_assigned}",
            f"Orders skipped (already in another group): "
            f"{orders_skipped_existing}",
            f"Orders unmatched (no SO with that ref): {len(orders_unmatched)}",
        ]
        if orders_unmatched:
            lines.append("")
            lines.append("Unmatched refs (first 20):")
            for r in orders_unmatched[:20]:
                lines.append(f"  {r}")

        self.summary = "\n".join(lines)
        return {
            "type": "ir.actions.act_window",
            "res_model": "wms.pick.group.import.wizard",
            "view_mode": "form",
            "res_id": self.id,
            "target": "new",
        }
